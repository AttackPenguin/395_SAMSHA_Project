from __future__ import annotations

import copy
import os
import pickle

import numpy as np
import openpyxl
import pandas as pd
from matplotlib import pyplot as plt

from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestClassifier
from sklearn.inspection import permutation_importance
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, classification_report, \
    ConfusionMatrixDisplay, PrecisionRecallDisplay, \
    RocCurveDisplay, roc_auc_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, OrdinalEncoder

import parameters as p

data_directory = p.DATA_DIRECTORY
performance_data_directory = \
    os.path.join(p.PERFORMANCE_DATA_DIRECTORY, 'Final Code')
classifier_directory = \
    os.path.join(p.CLASSIFIER_DIRECTORY, 'final_code')


def main():

    start_time = pd.Timestamp.now()
    print(f"Started at {start_time}")

    ############################################################################

    # destination = os.path.join(
    #     classifier_directory, "rf_prep_01"
    # )
    # with open(destination, 'rb') as file:
    #     preprocessing, rf_clf, feature_labels, X_train, X_test, y_train, y_test \
    #         = pickle.load(file)
    #
    # results = permutation_feature_importance(
    #     preprocessing, rf_clf, roc_auc_score, feature_labels,
    #     X_test, y_test, 20,
    #     file_name='Permutation Feature Importance, Preprocessor 01.pickle'
    # )

    destination = os.path.join(
        performance_data_directory, "Permutation Feature Importance, "
                                    "Preprocessor 01.pickle"
    )
    with open(destination, 'rb') as file:
        results = pickle.load(file)

    fig_pf_importance(results,
                      'Fig XX: Permutation Feature Importance, Preprocessor '
                      '01',
                      'Permutation Feature Importance, Preprocessor 01')

    ############################################################################

    # destination = os.path.join(
    #     classifier_directory, "rf_prep_01_opt"
    # )
    # with open(destination, 'rb') as file:
    #     preprocessing, rf_clf, feature_labels, X_train, X_test, y_train, y_test \
    #         = pickle.load(file)
    #
    # results = permutation_feature_importance(
    #     preprocessing, rf_clf, roc_auc_score, feature_labels,
    #     X_test, y_test, 20,
    #     file_name='Permutation Feature Importance, Preprocessor 01 '
    #               'Optimized.pickle'
    # )
    #
    destination = os.path.join(
        performance_data_directory, "Permutation Feature Importance, "
                                    "Preprocessor 01 Optimized.pickle"
    )
    with open(destination, 'rb') as file:
        results = pickle.load(file)

    fig_pf_importance(results,
                      'Fig XX: Permutation Feature Importance, Preprocessor '
                      '01 Optimized',
                      'Permutation Feature Importance, Preprocessor 01 '
                      'Optimized')

    ############################################################################

    # destination = os.path.join(
    #     classifier_directory, "rf_prep_02a"
    # )
    # with open(destination, 'rb') as file:
    #     preprocessing, rf_clf, feature_labels, X_train, X_test, y_train, y_test \
    #         = pickle.load(file)
    #
    # results = permutation_feature_importance(
    #     preprocessing, rf_clf, roc_auc_score, feature_labels,
    #     X_test, y_test, 20,
    #     file_name='Permutation Feature Importance, Preprocessor 02a.pickle'
    # )
    #
    destination = os.path.join(
        performance_data_directory, "Permutation Feature Importance, "
                                    "Preprocessor 02a.pickle"
    )
    with open(destination, 'rb') as file:
        results = pickle.load(file)

    fig_pf_importance(results,
                      'Fig XX: Permutation Feature Importance, Preprocessor '
                      '02a',
                      'Permutation Feature Importance, Preprocessor 02a')

    ############################################################################

    # destination = os.path.join(
    #     classifier_directory, "rf_prep_02b"
    # )
    # with open(destination, 'rb') as file:
    #     preprocessing, rf_clf, feature_labels, X_train, X_test, y_train, y_test \
    #         = pickle.load(file)
    #
    # results = permutation_feature_importance(
    #     preprocessing, rf_clf, roc_auc_score, feature_labels,
    #     X_test, y_test, 20,
    #     file_name='Permutation Feature Importance, Preprocessor 02b.pickle'
    # )
    #
    destination = os.path.join(
        performance_data_directory, "Permutation Feature Importance, "
                                    "Preprocessor 02b.pickle"
    )
    with open(destination, 'rb') as file:
        results = pickle.load(file)

    fig_pf_importance(results,
                      'Fig XX: Permutation Feature Importance, Preprocessor '
                      '02b',
                      'Permutation Feature Importance, Preprocessor 02b')

    ############################################################################

    end_time = pd.Timestamp.now()
    print(f"Finished at {end_time}")
    print(f"Run time {(end_time - start_time).total_seconds() / 60:.2f} "
          f"minutes.")


def fig_pf_importance(results: dict[str, list[float]],
                      title: str,
                      file_name: str):

    sorted_idx = np.array([np.mean(x) for x in results.values()]).argsort()
    importances = np.array([x for x in results.values()])

    fig: plt.Figure = plt.figure(figsize=[8.5, 11], dpi=400)
    ax: plt.Axes = fig.add_subplot()
    ax.boxplot(
        importances[sorted_idx].T, vert=False,
        labels=np.array(list(results.keys()))[sorted_idx]
    )
    # ax.set_title(title)
    ax.grid(axis='x')
    ax.set_xticks(
        [0.000, 0.010, 0.020, 0.030,
         0.040, 0.050, 0.060, 0.070]
    )
    ax.set_xlabel('ROC AUC Score Delta')
    fig.tight_layout()
    plt.show()
    destination = os.path.join(
        performance_data_directory,
        file_name
    )
    fig.savefig(destination)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, 'Feature')
    ws.cell(1, 2, 'Mean')
    ws.cell(1, 3, 'Std Dev')
    ws.cell(1, 4, 'Min')
    ws.cell(1, 5, 'Median')
    ws.cell(1, 6, 'Max')
    row = 2
    for result in results:
        ws.cell(row, 1, result)
        ws.cell(row, 2, np.mean(results[result]))
        ws.cell(row, 3, np.std(results[result]))
        ws.cell(row, 4, np.min(results[result]))
        ws.cell(row, 5, np.median(results[result]))
        ws.cell(row, 6, np.max(results[result]))
        row += 1

    wb.save(os.path.join(
        performance_data_directory, file_name+'.xlsx'
    ))



def permutation_feature_importance(preprocessor,
                                   classifier,
                                   scorer,
                                   features,
                                   X_score: pd.DataFrame,
                                   y_score: pd.Series,
                                   iterations: int = 10,
                                   file_name: str | bool = False):

    results = {feature: list() for feature in features}

    print(f"Generating initial baseline score...")
    t_X_score = preprocessor.transform(X_score)
    baseline_score = scorer(y_score, classifier.predict_proba(t_X_score)[:, 1])
    print(f"Initial baseline score: {baseline_score}")

    for feature in results.keys():
        print(f"\nBeginning analysis of feature {feature}...")
        for i in range(1, iterations+1):
            print(f"\tIteration {i} of {iterations}...")
            test_features = X_score.copy(True)
            test_features[feature] = \
                np.random.permutation(test_features[feature])
            t_test_features = preprocessor.transform(test_features)
            score = scorer(y_score,
                           classifier.predict_proba(t_test_features)[:, 1])
            results[feature].append(baseline_score-score)
        print(f"Analysis of feature {feature} complete.")
        print(f"Mean score delta = {np.mean(results[feature])}")

    if file_name:
        destination = os.path.join(
            performance_data_directory, file_name
        )
        with open(destination, 'wb') as file:
            pickle.dump(results, file)

    return results


if __name__ == '__main__':
    main()
