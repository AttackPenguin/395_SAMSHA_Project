import pickle
import os

import numpy as np
import openpyxl as openpyxl
import pandas as pd
from matplotlib import pyplot as plt
import seaborn as sns

from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, classification_report, \
    ConfusionMatrixDisplay, PrecisionRecallDisplay, \
    RocCurveDisplay, roc_auc_score

import parameters as p
import preprocessors
from parameters import FINAL_DATA

sns.set_theme(context='paper')

data_directory = p.DATA_DIRECTORY
performance_data_directory = \
    os.path.join(p.PERFORMANCE_DATA_DIRECTORY, 'Final Code',
                 'Subsampling Final')
classifier_directory = \
    os.path.join(p.CLASSIFIER_DIRECTORY, 'final_code')


def main():
    start_time = pd.Timestamp.now()
    print(f"Started at {start_time}")

    # target_file = os.path.join(data_directory, 'tedsd_puf_2019.csv')
    # raw_data = pd.read_csv(target_file)
    #
    # _, features = preprocessors.create_preprocessor_01_opt()
    # get_subsample(raw_data, features, 3)

    # size, auc = run_subsamples(37)
    # destination = os.path.join(performance_data_directory,
    #                            'subsampling_data.pickle')
    # with open(destination, 'wb') as file:
    #     pickle.dump((size, auc), file)

    destination = os.path.join(performance_data_directory,
                               'subsampling_data.pickle')
    with open(destination, 'rb') as file:
        data = pickle.load(file)

    subsampling_figure(data[0], data[1])

    end_time = pd.Timestamp.now()
    print(f"Finished at {end_time}")
    print(f"Run time {(end_time - start_time).total_seconds() / 60:.2f} "
          f"minutes.")


def subsampling_figure(samples: dict[int, int],
                       scores: dict[int, int]):
    fig: plt.Figure = plt.figure(figsize=[6.4, 4.6], dpi=400)
    ax: plt.Axes = fig.add_subplot()

    ax2: plt.Axes = ax.twinx()
    ax2.grid(False)

    # s1 = ax.scatter(list(scores.keys()),
    #                 list(scores.values()),
    #                 label='Scores', color='C0', s=3)
    # s2 = ax2.scatter(list(samples.keys()),
    #                  np.array(list(samples.values())) / 1000,
    #                  label='Samples', color='C1', s=3)

    s1, = ax.plot(list(scores.keys()),
                  list(scores.values()),
                  label='AUC After Re-Training\n'
                        'Classifier on Sub-Sample',
                  color='C0')
    s2, = ax2.plot(list(samples.keys()),
                   np.array(list(samples.values())) / 1000,
                   label='Number of Patients in\n'
                         'Subsample',
                   color='C1')

    ax.yaxis.label.set_color(s1.get_color())
    ax2.yaxis.label.set_color(s2.get_color())
    ax.tick_params(axis='y', colors=s1.get_color())
    ax2.tick_params(axis='y', colors=s2.get_color())

    ax.set_xlabel('Maximum Number of Missing Features')
    ax.set_ylabel('ROC AUC')
    ax2.set_ylabel('Patients in Data Set (thousands)')

    ax.legend(handles=[s1, s2], loc='lower right')

    fig.show()


def missing_feature_data_to_excel():
    """
    REPORT DATA - generates excel spreadsheet with missing data distributions.
    FINAL FORM - Output has been generated and this code does not need to be
    run again. Running it again will overwrite formatting to the spreadsheet
    performed post-generation.
    :return:
    """

    target_file = os.path.join(data_directory, 'tedsd_puf_2019.csv')
    raw_data = pd.read_csv(target_file)

    feature_sets = list()
    feature_sets.append(raw_data.columns)
    feature_sets.append(preprocessors.create_preprocessor_01()[1])
    feature_sets.append(preprocessors.create_preprocessor_01_opt()[1])
    feature_sets.append(preprocessors.create_preprocessor_02a()[1])
    feature_sets.append(preprocessors.create_preprocessor_02b()[1])

    set_names = ['raw data', '01', '01_opt', '02a', '02b']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Summary Data'

    ws.cell(1, 1, 'Preprocessor')
    ws.cell(1, 2, 'Number of Features')
    ws.cell(2, 1, 'Raw Data')
    ws.cell(2, 2, len(feature_sets[0]))
    ws.cell(3, 1, '01')
    ws.cell(3, 2, len(feature_sets[1]))
    ws.cell(4, 1, '01_opt')
    ws.cell(4, 2, len(feature_sets[2]))
    ws.cell(5, 1, '02a')
    ws.cell(5, 2, len(feature_sets[3]))
    ws.cell(6, 1, '02b')
    ws.cell(6, 2, len(feature_sets[4]))

    for i, feature_set in enumerate(feature_sets):
        x_vals, y_vals = get_missing_data_dist(raw_data, feature_set)
        ws = wb.create_sheet(set_names[i])
        ws.cell(1, 1, 'Features With Data')
        ws.cell(1, 2, 'Samples')
        row = 2
        for x, y in list(zip(x_vals, y_vals)):
            ws.cell(row, 1, x)
            ws.cell(row, 2, y)
            row += 1

    destination = os.path.join(
        FINAL_DATA,
        'Distribution of Missing Data in Feature Sets.xlsx'
    )
    wb.save(destination)


def get_missing_data_dist(raw_data: pd.DataFrame,
                          features: list[str]):
    x_vals = list()
    y_vals = list()

    for i in range(len(features)):
        sub_sample = raw_data[features].copy()
        sub_sample['sum'] = (sub_sample == -9).sum(axis=1)
        sub_sample = sub_sample[sub_sample['sum'] == i][features]
        x_vals.append(len(features) - i)
        y_vals.append(sub_sample.shape[0])
        print(f"{len(features) - i}\t{sub_sample.shape[0]}")

    fig: plt.Figure = plt.figure()
    ax: plt.Axes = fig.add_subplot()
    ax.plot(x_vals, y_vals)
    ax.set_xlabel('Number of Data Points Known')
    ax.set_ylabel('Number of Samples')
    fig.show()

    return x_vals, y_vals


def get_subsample(raw_data: pd.DataFrame,
                  features: list[str],
                  max_missing: int):
    raw_data = raw_data
    targets = raw_data['REASON']
    targets = np.array((targets == 3).astype(int))
    terminated = (targets == 1).sum()
    fx_term_raw = terminated / len(targets)

    sub_sample = raw_data[features + ['REASON']]
    sub_sample['sum'] = (sub_sample == -9).sum(axis=1)
    sub_sample = \
        sub_sample[sub_sample['sum'] <= max_missing][features + ['REASON']]

    terminated = len(sub_sample[sub_sample['REASON'] == 3])
    fx_term_ss = terminated / len(sub_sample)

    # Scenario where fx_term_ss is less than fx_term_raw: need to trim
    # non-terminated patients.
    if fx_term_ss < fx_term_raw:
        not_term = sub_sample[sub_sample['REASON'] != 3]
        term = sub_sample[sub_sample['REASON'] == 3]
        needed = int((len(term) / fx_term_raw) - len(term))
        not_term = not_term.sample(needed)
        sub_sample = pd.concat([term, not_term])

    # Scenario where fx_term_ss is greater than fx_term_raw: need to trim
    # terminated patients.
    if fx_term_ss > fx_term_raw:
        not_term = sub_sample[sub_sample['REASON'] != 3]
        term = sub_sample[sub_sample['REASON'] == 3]
        needed = int((fx_term_raw * len(not_term)) / (1 - fx_term_raw))
        term = term.sample(needed)
        sub_sample = pd.concat([term, not_term])

    print(f"Returning subsample of length {len(sub_sample)}")
    return sub_sample


def ss_random_forest(preprocessor: ColumnTransformer,
                     feature_labels: list[str],
                     X_train: pd.DataFrame,
                     X_test: pd.DataFrame,
                     y_train: pd.DataFrame,
                     y_test: pd.DataFrame,
                     trees: int = 1000):
    """
    :param trees:
    :param y_test:
    :param y_train:
    :param X_test:
    :param X_train:
    :param preprocessor:
    :param feature_labels:
    :param filename:
    :param filename_pretty:
    :return:
    """

    t_X_train = preprocessor.fit_transform(X_train)
    t_X_test = preprocessor.transform(X_test)

    print("Data loaded.")

    print("Starting training...")
    clf = RandomForestClassifier(
        n_jobs=10, verbose=1,
        n_estimators=trees, criterion='entropy',
        max_depth=None, max_features='sqrt',
        max_samples=None, class_weight='balanced_subsample'
    )

    clf.fit(t_X_train, y_train)

    score = 1 - roc_auc_score(y_test, clf.predict_proba(t_X_test)[:, 0])
    return score


def run_subsamples(max_missing: int):
    target_file = os.path.join(data_directory, 'tedsd_puf_2019.csv')
    raw_data = pd.read_csv(target_file)
    preprocessor, feature_labels = \
        preprocessors.create_preprocessor_01_opt()

    subsample_size = dict()
    roc_auc = dict()

    for missing in range(max_missing):
        print(f"Running missing={missing}...")

        subsample = get_subsample(raw_data.copy(), feature_labels, missing)
        subsample_size[missing] = len(subsample)

        targets = subsample.pop('REASON')
        targets = np.array((targets == 3).astype(int))
        X_train, X_test, y_train, y_test = \
            train_test_split(subsample, targets,
                             test_size=0.1, stratify=targets)

        score = ss_random_forest(preprocessor,
                                 feature_labels,
                                 X_train, X_test, y_train, y_test,
                                 trees=1000)

        roc_auc[missing] = score
        print(f"\tSubsample Size:", len(subsample))
        print(f"\troc_auc:", score)

    return subsample_size, roc_auc


if __name__ == '__main__':
    main()
